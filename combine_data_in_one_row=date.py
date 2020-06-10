from pykrx import stock
import numpy as np
import openpyxl as op
import pandas as pd
from collections import OrderedDict

tickers = stock.get_market_ticker_list() # collect stock num
stock_num = np.array(tickers)
company_list = OrderedDict()
company_list_num = list() # key : 순서 번호 / value :  기업 종목번호_기업명 을
company_cost_ratio = OrderedDict()
date_len = 0

def read_combine_excel():
    wb = op.Workbook()
    wb = op.load_workbook("./data/combine_data.xlsx")
    sheet1 = wb["Sheet"]
    for idx,val in enumerate(sheet1):
        tmp2 = list()
        if idx == 0 : # 첫째 줄.. 시간
            for i,date in enumerate(val):
                if i<=1: # 첫번째줄. 시간 행
                    continue
                tmp2.append(date.value)
            date_len = len(tmp2)
            company_list['date'] = tmp2
        else:
            for i,j in enumerate(val) :
                #print(j.value ,end= " / ")
                if i <=1 :
                    continue
                tmp2.append(j.value)
            company_list[val[0].value + "_" + val[1].value] = tmp2
            company_list_num.append(val[0].value + "_" + val[1].value)
    return date_len
        #print()
    #wb.close()

def make_ratio_excel():
    wf = op.Workbook()
    wb = wf.active
    wb['A1'] = "종목코드";
    wb['B1'] = "회사명"
    for i,val in enumerate(company_list['date']):
        if i ==0:
            continue
        wb.cell(row=1,column=i+2).value = val

    for idx,(key,val) in enumerate(company_cost_ratio.items()):
        data_ = key.split("_")
        wb['A' + str(idx + 2)] = data_[0]
        wb['B' + str(idx + 2)] = data_[1]
        # print(val)
        # input()
        for i,ratio in enumerate(val):
            wb.cell(row=idx+2,column=i+3).value = ratio
    wf.save("./data/combine_data_ratio.xlsx")

date_len = read_combine_excel()

for key,val in company_list.items():
    if key == "date":
        continue
    new_cost = -1
    old_cost = -1
    ratio = list()
    for cost in val:
        new_cost = cost
        if old_cost ==-1:
            old_cost = new_cost
            continue
        ratio.append(round(((new_cost-old_cost)  / new_cost) * 100,4))
        old_cost = new_cost
    company_cost_ratio[key]=ratio
# print(len(company_list['date']))
# input("CHECK")
date_ratio_list = OrderedDict()
#date_ratio_list는 key 값은 시간 value 값은 100개 종목 ratio 순서대로 되어있음


for idx, (key,val) in enumerate(company_cost_ratio.items() ):
    # if idx<10:
    #     print(idx , " + ",key," : ",end=" ")
    #     print(val)
    for i,ratio in enumerate(val):
        tmp = list()
        index = company_list['date'][i+1]

        if date_ratio_list.get(index):
            tmp = date_ratio_list[index]
            tmp.append((ratio,len(date_ratio_list[index])))
            date_ratio_list[index] = tmp
        else:
            tmp.append((ratio,0))
            date_ratio_list[index]=tmp

for idx, (key, val) in enumerate(date_ratio_list.items()):
    if idx==0:
        print(key,val)

make_ratio_excel()

print("-"*100)
for idx, (key, val) in enumerate(date_ratio_list.items()):
    date_ratio_list[key]=sorted(date_ratio_list[key],key=lambda x : x[0],reverse=True)


### top5 bottom5 excel 파일 만들기
wf = op.Workbook()
wb = wf.active
wb['A1'] = "날짜";
wb['B1'] = "회사명"

for idx, (key, val) in enumerate(date_ratio_list.items()):
    print(key)
    wb.cell(row=idx+2, column=1).value = key
#    input("PAUSE")
    for i,ratio in enumerate(date_ratio_list[key]):
        if i<5:
            if i ==0:
                print("상위 5개")
            company_name = company_list_num[ratio[1]]
            print(company_name," / ratio : ",ratio[0])
            company_name=company_name.split("_")
            #날짜 열에는 회사명 비율 열에는 등락율
            wb.cell(row=idx+2, column=i+2).value = company_name[1]
            #wb.cell(row=i+3, column=2 * idx + 3).value = ratio[0]
    print("-"*100)

#input("STOP")
wf.save("./data/combine_data_ratio_row=date.csv")
wf.save("./data/combine_data_ratio_row=date.xlsx")
    # if idx < 5:
    #     print(date_ratio_list[key])
    #     continue
    # print(idx , " + ",key," : ",end=" ")
    # print(val)





