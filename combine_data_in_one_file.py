from pykrx import stock
import numpy as np
import openpyxl as op
import pandas as pd


company_list=list()
def stock_name_num():
    wb = op.Workbook()
    wb = op.load_workbook("data/stock_list_with_name&num.xlsx")
    sheet1 = wb["Sheet"]
    for idx,val in enumerate(sheet1):
        if idx == 0:
            continue
        company_list.append([val[0].value,val[1].value])
        #(종목명,종목번호)
    wb.close()

def add_zero(num):
    num = str(num)
    num = "0" * (6 - len(num)) + num
    return str(num)


stock_name_num()
wf = op.Workbook()
wb = wf.active
wb['A1'] = "종목코드";
wb['B1'] = "회사명"
for idx,val in enumerate(company_list):
    df = stock.get_market_ohlcv_by_date("20170101", "20200406", val[1])
    wb['A'+str(idx+2)]= val[1];wb['B'+str(idx+2)]=val[0]
    print(val[1])
    print(val)

    date_ = df.index
    end_costs = df.values
    date_idx = list()
    end_cost_values = list()
    # time index
    if idx == 0:
        for i in date_:
            date_idx_tmp = str(i)
            date_idx.append(date_idx_tmp.split(" ")[0])
        for i,date_ in enumerate(date_idx):
            wb.cell(row=1,column=i+3).value = date_
        # end_index
    for i in end_costs:
        # stock.get_market_ticker_name("종목번호")
        # print(stock.get_market_ticker_name(add_zero(val[1])))
        end_cost_values.append(i[3])

    new_cost = -1
    old_cost = -1
    ratio = list()
    for i,cost_ in enumerate(end_cost_values):
        new_cost = cost_
        if old_cost != -1:
            ratio.append(round((new_cost - old_cost)/old_cost * 100,4))
        old_cost = new_cost
        wb.cell(row=idx+2, column=i+3).value = cost_


wf.save("./data/combine_data.xlsx")


