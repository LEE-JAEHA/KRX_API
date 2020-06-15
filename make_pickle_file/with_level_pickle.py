from pykrx import stock
import numpy as np
import openpyxl as op
import pandas as pd
import pickle
import os.path

def read_indicator_and_write(wb):
    """
    wb "../data/final/normalized/combine_level_indicator_MinMax_sheet_date.xlsx" 에 시트 이름은 날짜 시트 내용은 해당 일 기업 별 보조 지표 데이터
    indicator는
    indicator[날짜][기업명] 에 대한 보조 지표 데이터들이 있다.
    """
    name_=""
    indicator=dict()
    for sheet in wb:
        date_ = sheet.title #시트 이름은 날짜 이름
        print(date_)
        for idx,val in enumerate(sheet):
            if idx == 0: # 첫번째 줄 index
                continue
            tmp = list()
            for v in val:
                tmp.append(v.value)
            company = tmp[1]
            tmp.pop(0)
            print(company)
            new_ = {company: tmp}
            if date_ in indicator:
                tmp2 = indicator[date_]
                indicator[date_].update(new_)
            else:
                indicator[date_ ]=new_
    file = open("../data/indicator/with_level_indicator","wb")
    pickle.dump(indicator,file)
    file.close()

    return indicator

wb2 = op.Workbook()
if os.path.exists("../data/indicator/with_level_indicator"):
    file = open("../data/indicator/with_level_indicator", "rb")
    indicator = pickle.load(file)
else:
    wb = op.Workbook()
    print("OK")
    wb = op.load_workbook("../data/final/normalized/combine_level_indicator_MinMax_sheet_date.xlsx")
    indicator = dict()
    indicator = read_indicator_and_write(wb)
    print("FINISH")



#make_indicator_file_sheet_name_date(company_list,total_data,wb2)


