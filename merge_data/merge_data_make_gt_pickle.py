from pykrx import stock
import numpy as np
import openpyxl as op
import pandas as pd
import pickle
import os.path

def stock_name_num():
    wb = op.Workbook()
    wb = op.load_workbook("../data/modify_data.xlsx")
    sheet1 = wb["Sheet1"]
    for idx,val in enumerate(sheet1):
        if idx == 0:
            continue
        company_list.append([val[0].value,val[1].value])
        #(종목명,종목번호)
    wb.close()
    return company_list


def read_indicator_and_write(company_list,wb2):
    """
    wb ./data/indicator/stock_data_indicator.xlsx에 있는 데이터를 읽어와
    wb2 ./data/indicator/combine_indicator.xlsx 에 시트 이름은 날짜 시트 내용은 해당 일 기업 별 보조 지표 데이터
    indicator는
    indicator[날짜][기업명] 에 대한 보조 지표 데이터들이 있다.
    """
    print("write indicator file")
    name_=""
    menu = ['회사명', 'close', 'diff', 'open', 'high', 'low', 'volume', 'MMS_MA20P', 'MMS_MA60P', 'MMS_MA120P', 'MMS_ATR', 'MMS_slowk', 'MMS_slowd', 'MMS_MOM', 'MMS_RSI', 'MMS_ADX', 'MMS_macd', 'MMS_macdsignal', 'MMS_macdhist', 'MMS_aroondown', 'MMS_aroonup', 'MMS_VAR', 'MMS_WILLR', 'search', 'combine']
    indicator=dict()
    for sheet in wb2:
        date_ = sheet.title
        for idx,val in enumerate(sheet):
            if idx == 0:#메뉴줄
                continue
            tmp = list()
            for v in val:
                tmp.append(v.value)
            tmp.append(False)
            cp_name = tmp.pop(0)
            new_ = {cp_name:tmp}
            if date_ in indicator:
                tmp2 = indicator[date_]
                indicator[date_].update(new_)
            else:
                indicator[date_]=new_
    return indicator


def top5_list(indicator):
    wb = op.Workbook()
    wb = op.load_workbook("../data/top5_dataset_indicator_0604.xlsx")
    sheet = wb["TOP5"]
    for idx,val in enumerate(sheet):
        if idx == 0:
            continue # 첫번째 줄
        for i,v in enumerate(val):
            if i ==2 :
                cp_name = v.value
            elif i ==3:
                date_ = str(v.value).split(" ")[0]
                break
        indicator[date_][cp_name][-1]=True
    file = open("../data/indicator/indicator_dict3", "wb")
    pickle.dump(indicator, file)
    file.close()

def increase_list(indicator):
    for date_,cp_data in indicator.items():
        #date_ 는 key / cp_data는 date_에 해당하는 기업들의 보조지표 얘는 key는 기업명 value는 보조지표
        for cp_name,indi in cp_data.items():
            indicator[date_][cp_name][-1] = indi[-2] >  0.25
    file = open("../data/indicator/indicator_dict3", "wb")
    pickle.dump(indicator, file)
    file.close()
    return indicator



def make_excel_file(indicator):
    print("MAKE EXCEL FILE")
    wb2 = op.Workbook()
    wb2 = op.load_workbook("../data/final/merged (3).xlsx")
    for sheet in wb2:
        date_ = sheet.title
        #wb.cell(row=idx+2, column=1).value = key
        #sheet.cell(row=idx+1,column=)
        for idx,val in enumerate(sheet):
            if idx ==0:
                sheet.cell(row=idx + 1, column=len(val)+1).value= "TOP5"
                continue
            #print(val[0].value) 회사명
            #print(indicator[date_][val[0].value][-1]) TOP5인지 아닌지
            sheet.cell(row=idx + 1, column=len(val)+1).value = indicator[date_][val[0].value][-1]
    wb2.close()
    wb2.save("../data/final/tmp.xlsx")






company_list=list()
company_list = stock_name_num()
if os.path.exists("../data/indicator/indicator_dict3"):
    print("HAVE PICKLE")
    file = open("../data/indicator/indicator_dict3", "rb")
    indicator = pickle.load(file)
    #top5_list(indicator)
    indicator = increase_list(indicator)
    #make_excel_file(indicator)
else:
    print("pickle make")
    wb2 = op.Workbook()
    wb2 = op.load_workbook("../data/final/merged (3).xlsx")
    indicator = dict()
    indicator = read_indicator_and_write(company_list,wb2)
    indicator = increase_list(indicator)
    #top5_list(indicator)
    make_excel_file(indicator)
    # wb2.remove(wb2["Sheet"])
    # wb2.save("./data/indicator/combine_indicator.xlsx")
