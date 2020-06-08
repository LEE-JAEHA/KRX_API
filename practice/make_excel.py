import openpyxl as op

#
# wb = op.Workbook()
# wb = op.load_workbook("../data/indicator/stock_data_indicator.xlsx")
#
# for i in wb:# i는 sheet 하나씩
#     for j in i : # j는 i sheet에서 한줄 씩
#         for k in j:
#             print(k.value, end =" ")
#
#
# wb.active
# print("HI")
#

from tqdm import tqdm_notebook
from tqdm import trange
from tqdm import tqdm
import time
for i in tqdm(range(10)):
    time.sleep(0.1)

for i in trange(10):
    time.sleep(0.1)
    print(10)



#
#
#
# wb = op.Workbook()
#
# sheet1 = wb.active
# sheet2 = wb.create_sheet("두번째 시트")
# sheet2['B2'] = 'b2'
# sheet2.cell(row=3, column=3).value = '3, 3'
# sheet2.append([1, 2, 3, 4, 5])
# wb.close()
#
# wb.save('../data/practice/test.xlsx')

 
 
