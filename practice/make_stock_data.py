from pykrx import stock
import numpy as np
import openpyxl as op
import pandas as pd


company_list=list()
def stock_name_num():
    wb = op.Workbook()
    wb = op.load_workbook("../data/stock_list_with_name&num.xlsx")
    sheet1 = wb["Sheet"]
    for idx,val in enumerate(sheet1):
        if idx == 0:
            continue
        company_list.append([val[0].value,val[1].value])
        #(종목명,종목번호)
    wb.close()

def add_zero(num):
    num = str(num)
    num = "0" * (6 - len(num)) + num
    return str(num)


stock_name_num()
for idx,val in enumerate(company_list):
    wf = op.Workbook()
    wb = wf.active
    df = stock.get_market_ohlcv_by_date("20160101", "20200406", val[1])
    wb['A1'] = val[0];
    wb['A2'] = "날짜";
    wb['B2'] = "종가"
    print(val[1])
    date_ = df.index
    end_costs = df.values
    date_idx = list()
    end_cost_values = list()


    # time index
    for i in date_:
        date_idx_tmp = str(i)
        date_idx.append(date_idx_tmp.split(" ")[0])
    for i,date_ in enumerate(date_idx):
        wb.cell(row=i+3,column=1).value = date_
    # end_index
    for i in end_costs:
        # stock.get_market_ticker_name("종목번호")
        # print(stock.get_market_ticker_name(add_zero(val[1])))
        end_cost_values.append(i[3])


    for i,cost_ in enumerate(end_cost_values):
        wb.cell(row=i +3, column=2).value = cost_

    file_name = val[1] + "_" + val[0]
    wf.save("./data/stock_data/" + file_name + ".xlsx")


