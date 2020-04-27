import openpyxl as op
from pykrx import stock
import numpy as np
import pandas as pd

tickers = stock.get_market_ticker_list() # collect stock num
stock_num = np.array(tickers)
company_list=list()
def stock_name():
    wb = op.Workbook()
    wb = op.load_workbook("./data/stock_list.xlsx")
    sheet1 = wb["Sheet1"]
    for idx,val in enumerate(sheet1):
        if idx == 0:
            continue
        company_list.append(val[0].value)
    wb.close()
def add_zero(num):
    num = str(num)
    num = "0"*(6-len(num))+num
    return num


stock_name()
new_com_list=list()
for i in tickers:
    #print(i)
    num_ = stock.get_market_ticker_name(i)
    if num_ in company_list:
        new_com_list.append([num_,i])

print(new_com_list)


wf= op.Workbook()
wb2 = wf.active
wb2['A1']="종목 번호";wb2['B1']="종목 명"
for i in range(len(new_com_list)):
    wb2.cell(row=i+2,column=1).value = new_com_list[i][0]
    wb2.cell(row=i+2, column=2).value = new_com_list[i][1]
wf.save("./data/stock_list_with_name&num.xlsx")



