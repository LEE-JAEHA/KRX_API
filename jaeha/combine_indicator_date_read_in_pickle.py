from pykrx import stock
import numpy as np
import openpyxl as op
import pandas as pd
import pickle


def stock_name_num(path):
    company_list=list()
    wb = op.Workbook()
    wb = op.load_workbook("./data/modify_data.xlsx")
    sheet1 = wb["Sheet1"]
    for idx,val in enumerate(sheet1):
        if idx == 0:
            continue
        company_list.append([val[0].value,val[1].value])
        #(종목명,종목번호)
    wb.close()
    return company_list


def make_indicator_file_sheet_name_date(company_list,indicator):
    """
    :param company_list: 회사명 리스트 company_list[0][0] 종목명 company_list[0][1] 종목 코드
    :param indicator: indicator[날짜][기업명] 에 대한 보조 지표 데이터들이 있다.
    :param wb: ./data/indicator/combine_indicator.xlsx 에 시트 이름은 날짜 시트 내용은 해당 일 기업 별 보조 지표 데이터
    :return:
    """

    path = "./data/indicator/combine_indicator.xlsx"
    wb = op.Workbook()
    wb = op.load_workbook(path)

    # for sheet in wb:
    #     for idx,val in enumerate(sheet):
    #         if idx ==0:
    #             continue
    #         print(val[1].value)
    for keys,values in indicator.items():#keys는 날짜
        print(keys, "  FINISH")
        if keys in wb.sheetnames:
            for idx, val in enumerate(wb[keys]):  # date_에 하나씩 기록하기 / val는 한줄 씩 / 이 파일에는 이미 종목명 회사 적혀 있음 보조 지표 메뉴까지
                if idx == 0:  # 메뉴 부분
                    continue
                name_ = val[1].value  # ex_) CJ
                code_ = val[0].value
                # indicator[date_][name_] date_ 에 name_ 의 보조지표
                comnpany_indicator_data = indicator[keys][name_]
                for i, data_ in enumerate(comnpany_indicator_data):
                    wb[keys].cell(row=idx + 1, column=i + 3).value = data_
    wb.close()
    wb.save("./data/indicator/333.xlsx")


    """
    for keys,values in indicator.items():#keys는 날짜
        print(keys, "  FINISH")
        for k,v in indicator[keys].items(): # k는 기업명 v는 보조지표 값들
            #print(v[0])
            date_ = str(v[0]).split(" ")[0]
            # a =  wb.sheetnames
            # #print(type(a))
            # if date_ in a:
            #     print("HI")
            if date_ in wb.sheetnames:
                for idx,val in enumerate(wb[date_]): # date_에 하나씩 기록하기 / val는 한줄 씩 / 이 파일에는 이미 종목명 회사 적혀 있음 보조 지표 메뉴까지
                    if idx ==0: # 메뉴 부분
                        continue
                    name_ = val[1].value #ex_) CJ
                    code_ = val[0].value
                    #indicator[date_][name_] date_ 에 name_ 의 보조지표
                    comnpany_indicator_data = indicator[date_][name_]
                    for i,data_ in enumerate(comnpany_indicator_data):
                        wb[date_].cell(row = idx+1,column=i+3).value = data_
    wb.close()
    wb.save("./data/indicator/333.xlsx")
    """







file = open("./data/indicator/indicator_dict","rb")
indicator = pickle.load(file)
print("HI")
company_list= stock_name_num("./data/modify_data.xlsx")
print("FINISH")
make_indicator_file_sheet_name_date(company_list,indicator)
print("FINISH2")