from pykrx import stock
import numpy as np
import openpyxl as op
import pandas as pd
import pickle
import os.path

company_list=list()
def tmp_stock_list():
    """
    stock list 가 잘못되어 수정전 임시 파일로 기업 리스트 읽어옴
    :return:
    """
    wb = op.Workbook()
    wb = op.load_workbook("data/indicator/tmp_stock_list.xlsx")
    sheet1 = wb["Sheet"]
    for idx, val in enumerate(sheet1):
        if idx == 0:
            continue
        company_list.append([val[0].value, val[1].value])
        # (종목명,종목번호)
    wb.close()
def stock_name_num():
    wb = op.Workbook()
    wb = op.load_workbook("./data/modify_data.xlsx")
    sheet1 = wb["Sheet1"]
    for idx,val in enumerate(sheet1):
        if idx == 0:
            continue
        company_list.append([val[0].value,val[1].value])
        #(종목명,종목번호)
    wb.close()
    return company_list

def indicator_menu_list(company_list,wb,wb2):
    """
    :param company_list: 기업명
    :param wb: #wb => "./data/indicator/stock_data_indicator.xlsx" 보조지표 전체 리스트
    :param wb2: 내가 새롭게 만들 시트 / 날짜별 기업 지표
    이 함수가 끝나면 시트 이름은 시간 각 시트 내용은 보조 지표에 대한 파일이 만들어진다.
    """

    sheet1 = wb[company_list[1][0]]
    print(company_list[1][1])
    menu = list() # 보조지표 종류
    for idx, val in enumerate(sheet1):
        if idx == 0:
            for v in val:
                menu.append(v.value)
        if idx == 1:
            for v in val:
                sheet_name_date = v.value
                break
            break
        # (종목명,종목번호)
    date_list=list()
    for sheet_ in wb:
        for i,v in enumerate(sheet_):
            if i ==0:
                continue
            print(v[0].value)
            tmp = str((v[0].value)).split(" ")
            # tmp = str(sheet_.cell(row=2,column=1).value).split(" ")
            date_list.append(tmp[0])
        break

    for new_sheet in date_list:
        write_sheet = wb2.create_sheet(new_sheet)
        write_sheet['A1'] = "종목코드"
        write_sheet['B1'] = "회사명"
        for idx, val in enumerate(menu):
            write_sheet.cell(row=1, column=idx + 3).value = val
        for idx, val in enumerate(company_list):
            write_sheet['A' + str(idx + 2)] = val[1]
            write_sheet['B' + str(idx + 2)] = val[0]


def read_indicator_and_write(company_list,wb,wb2):
    """
    wb ./data/indicator/stock_data_indicator.xlsx에 있는 데이터를 읽어와
    wb2 ./data/indicator/combine_indicator.xlsx 에 시트 이름은 날짜 시트 내용은 해당 일 기업 별 보조 지표 데이터
    indicator는
    indicator[날짜][기업명] 에 대한 보조 지표 데이터들이 있다.
    """
    name_=""
    indicator=dict()
    for sheet in wb:
        company = sheet.title
        for idx,val in enumerate(sheet):
            if idx == 0:
                continue
            date_ = str(val[0].value).split(" ")[0]
            tmp = list()
            for v in val:
                tmp.append(v.value)
            tmp.append(False)
            new_ = {company: tmp}
            if date_ in indicator:
                tmp2 = indicator[date_]
                indicator[date_].update(new_)
            else:
                indicator[date_ ]=new_
    file = open("./data/indicator/indicator_dict2","wb")
    pickle.dump(indicator,file)
    file.close()

    return indicator

def make_indicator_file_sheet_name_date(company_list,indicator,wb2):
    """
    :param company_list: 회사명 리스트 company_list[0][0] 종목명 company_list[0][1] 종목 코드
    :param indicator: indicator[날짜][기업명] 에 대한 보조 지표 데이터들이 있다.
    :param wb2: ./data/indicator/combine_indicator.xlsx 에 시트 이름은 날짜 시트 내용은 해당 일 기업 별 보조 지표 데이터
    :return:
    """
    for keys,values in indicator.items():#keys는 날짜
        for k,v in indicator[keys].items(): # k는 기업명 v는 보조지표 값들
            print(v[0])
            date_ = str(v[0]).split(" ")[0]
            #wb2[date_].cell()
            for k in wb2[date_]:
                print(k)





company_list = stock_name_num()
#tmp_stock_list() # TODO :수정된 기업 리스트로 다시 해야 함
wb2 = op.Workbook()
if os.path.exists("./data/indicator/indicator_dict2"):
    file = open("./data/indicator/indicator_dict2", "rb")
    indicator = pickle.load(file)
else:
    wb = op.Workbook()
    wb = op.load_workbook("./data/indicator/stock_data_indicator_level_0604_new.xlsx")

    indicator_menu_list(company_list,wb,wb2)
    indicator = dict()
    indicator = read_indicator_and_write(company_list, wb, wb2)
    wb2.remove(wb2["Sheet"])
    wb2.close()
    wb2.save("./data/indicator/combine_indicator.xlsx")


#make_indicator_file_sheet_name_date(company_list,total_data,wb2)


