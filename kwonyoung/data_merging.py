from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np



load_wb = load_workbook("stock_data_indicator_normalize_0603_new.xlsx", data_only=True)
sheet_list = load_wb.get_sheet_names()
print(f"sheet_list is {sheet_list}")
print(f"len(sheet_list) is {len(sheet_list)}")

data = []
for sheet in sheet_list:
    #data_cell = load_wb[sheet]['B1:AA839'] #all
    #data_cell = load_wb[sheet]['L1':'AA839'] #from MA20P
    data_cell = load_wb[sheet]['B1:W839'] #normalize all
    #data_cell = load_wb[sheet]['H1':'W839'] #normalize from MA20P
    tmp_sheet = []
    for cell in data_cell:
        tmp_row = []
        for element in cell:
            tmp_row.append(element.value)
        tmp_sheet.append(tmp_row)
    data.append(tmp_sheet)

print(f'data shape is {np.array(data).shape}')


load_ws = load_wb[sheet_list[0]]
print(load_ws['D2'].value)
get_cells = load_ws['C2':'M101']



date_cell = load_ws['A']

date = []

for cell in date_cell:
    date.append(cell.value)

date = date[1:]

#date = date.tolist()

print(f"date is {date}")
print(f"len(date) is {len(date)}")

load_wb_naver = load_workbook("naver_search_trend_0605.xlsx", data_only=True)
load_wb_combine = load_workbook("combine_data_ratio.xlsx", data_only=True)

load_ws_naver = load_wb_naver['시트1']
load_ws_combine = load_wb_combine['Sheet']

date_naver = load_ws_naver['1']
date_combine = load_ws_combine['1']

date_naver_str = []
for i in range(len(date_naver)-1):
    date_naver_str.append(date_naver[i+1].value)

date_naver_str = [x.strftime("%Y-%m-%d") for x in date_naver_str]

date_combine_str = []
for i in range(len(date_combine)-1):
    date_combine_str.append(date_combine[i+1].value)

print(f'date_naver_str is {date_naver_str}')
print(f'date_combine_str is {date_combine_str}')

print(f'type(date_naver_str) is {type(date_naver_str)}')
print(f'type(date_combine_str) is {type(date_combine_str)}')

get_cells_naver = load_ws_naver['B2:ATK101']
get_cells_combine = load_ws_naver['B2:ADS101']

naver_dict = dict()
combine_dict = dict()

print(f'len(date_naver_str) is {len(date_naver_str)}')
print(f'len(date_combine_str) is {len(date_combine_str)}')
print(f'len(get_cells_naver) is {len(get_cells_naver)}')
print(f'np.array(get_cells_naver).shape is {np.array(get_cells_naver).shape}')
print(f'len(get_cells_combine) is {len(get_cells_combine)}')
print(f'np.array(get_cells_combine).shape is {np.array(get_cells_combine).shape}')

for i in range(min(len(date_naver_str), np.array(get_cells_naver).shape[1])):
    tmp = [x.value for x in np.array(get_cells_naver)[:,i]]
    naver_dict[date_naver_str[i]] = tmp

for i in range(len(date_combine_str)):
    combine_dict[date_combine_str[i]] = [x.value for x in np.array(get_cells_combine)[:,i]]

print(f'naver_dict.keys() is {naver_dict.keys()}')
print(f'combine_dict.keys() is {combine_dict.keys()}')


write_wb = Workbook()
 

for i in range(min(len(date), len(date_naver), len(date_combine))-1):
    date_str = date[i+1].strftime("%Y-%m-%d")
    write_ws = write_wb.create_sheet(date_str)
    sheet_append = []
    for j in range(len(data)):
        sheet_append.append(data[j][i+1])
    
    #print(f'data[0][0] is {data[0][0]}')
    write_ws.append(['회사명'] + data[0][0] + ["search", "combine"])
    for i, row in enumerate(sheet_append):
        write_ws.append([sheet_list[i]] + row + [naver_dict[date_str][i], combine_dict[date_str][i]])



write_wb.save('merged.xlsx')