from sklearn.cluster import DBSCAN
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np

X = np.array([[1, 2, 1], [2, 2, 1], [2, 3, 1], [8, 7, 1], [8, 8, 1], [25, 80, 1]])
clustering = DBSCAN(eps=3, min_samples=2).fit(X)
print(clustering.labels_)
clustering

# load_wb_indicator = load_workbook("indicator.xlsx", data_only=True)
# load_wb_naver = load_workbook("naver_search_trend_0605.xlsx", data_only=True)
# load_wb_combine = load_workbook("combine_data_ratio.xlsx", data_only=True)

load_wb = load_workbook("merged_with_0.25up.xlsx", data_only=True)

sheet_list = load_wb.get_sheet_names()

all_values = []
increase_values = []

for sheet in sheet_list:

    load_ws = load_wb[sheet]
    get_cells = load_ws['H2':'X101']
    get_cells_increase = load_ws['Z']

    all_values_sheet = []
    for row in get_cells:
        row_value = []
        for cell in row:
            if cell.value == None:
                #print(cell.value)
                print(f'{cell} is None!')
                #print(row)
                row_value.append(0)

            else:
                row_value.append(cell.value)
        all_values_sheet.append(row_value)

    all_values += all_values_sheet

    increase_values_sheet = []
    for row in get_cells_increase:
        if row.value == None:
            print(row.value)
            print(f'{row} is None!')
            row_value.append(0)
        elif row.value == True:
            increase_values_sheet.append(1)
        elif row.value == False:
            increase_values_sheet.append(0)

    increase_values += increase_values_sheet

    assert len(increase_values_sheet) == 100, print(f'len(increase_values) is {len(increase_values_sheet)}')
    # print(f'increase_values_sheet is {increase_values_sheet}')

print(f'all_values[0] is {all_values[0]}')
print(f'len(all_values) is {len(all_values)}')
print(f'len(all_values[0]) is {len(all_values[0])}')
print(f'np.array(all_values).shape is {np.array(all_values).shape}')

# clustering = DBSCAN(eps=105, min_samples=3).fit(all_values)
# print(f'eps = {105} min_samples = {3}\n{clustering.labels_}\n')

train_size = 70000
test_size = 30000

all_values = all_values[:-100]
increase_values = increase_values[100:]

all_values = all_values[10000:]
increase_values = increase_values[10000:]

all_values = all_values[:train_size]
increase_values = increase_values[:train_size]

eps_sam = []
eps_cand = [70, 75, 80, 85, 90, 95, 100, 105, 110, 115, 120, 125, 130]
min_sam_cand = [2, 3, 4, 5, 6, 7, 8, 9, 10]
cnt = 0
import copy

all_labels = []
all_labels_original = []
all_cluster_label = []
for i, eps_val in enumerate(eps_cand):
    eps_ = []
    for j, min_sam_val in enumerate(min_sam_cand):
        clustering = DBSCAN(eps=eps_val, min_samples=min_sam_val).fit(all_values)
        print(f'eps = {i} min_samples = {j}\n{clustering.labels_}\n')

        gt_dict = {}
        pred_dict = {}
        for gt_each in increase_values[:-test_size]:
            gt_dict[str(gt_each)] = gt_dict.get(str(gt_each), 0) + 1
        for pred_each in clustering.labels_.tolist()[:-test_size]:
            pred_dict[str(pred_each)] = pred_dict.get(str(pred_each), 0) + 1

        gt_res = sorted(gt_dict.items(), key=(lambda x: x[1]), reverse=True)
        pred_res = sorted(pred_dict.items(), key=(lambda x: x[1]), reverse=True)

        print(f'pred_res before is {pred_res}')

        if len(pred_res) > 1 and int(pred_res[1][0]) == -1:
            pred_res_tmp = pred_res[1]
            pred_res.pop(1)
            pred_res = pred_res + [pred_res_tmp]

        print(f'gt_res is {gt_res}')
        print(f'int(gt_res[0][0]) is {int(gt_res[0][0])}')
        print(f'int(gt_res[1][0]) is {int(gt_res[1][0])}')
        print(f'pred_res is {pred_res}')


        ################### voting system #######################################

        cluster_label_dict = {}
        cluster_labels = clustering.labels_
        print(f'cluster_labels is {cluster_labels}')
        print(f'pred_dict.keys() is {pred_dict.keys()}')
        for cluster_label in pred_dict.keys():
            cluster_indexes = np.where(cluster_labels==int(cluster_label))
            #print(f'cluster_indexes is {cluster_indexes}')
            count_dict = {}
            for cluster_index in cluster_indexes[0]:
                #print(f'cluster_index is {cluster_index}')
                #print(f'increase_values[cluster_index] is {increase_values[cluster_index]}')
                count_dict[increase_values[cluster_index]] = count_dict.get(increase_values[cluster_index], 0) + 1
            print(f'count_dict is {count_dict}')
            count_res = sorted(count_dict.items(), key=(lambda x: x[1]), reverse=True)
            print(f'count_res is {count_res}')
            print(f'int(cluster_label) is {int(cluster_label)}')
            cluster_label_dict[int(cluster_label)] = count_res[0][0]

        print(f'cluster_label_dict is {cluster_label_dict}')
        all_cluster_label.append(cluster_label_dict)

        labels = []
        labels_original = copy.deepcopy(clustering.labels_.tolist())

        for label in clustering.labels_:
            #print(f'label is {label}')
            if label in cluster_label_dict.keys():
                labels.append(cluster_label_dict[label])
            else:
                labels.append(0)

        #print(f'labels is {labels}')

        ########################################################################3


        """
        labels = []
        labels_original = copy.deepcopy(clustering.labels_.tolist())
        for label in clustering.labels_:
            if label == int(pred_res[0][0]):
                labels.append(int(gt_res[0][0]))
            elif label == int(pred_res[1][0]):
                labels.append(int(gt_res[1][0]))
            elif label == -1:
                labels.append(int(gt_res[0][0]))
            else:
                labels.append(int(gt_res[1][0]))
        """

        """
        x = abs(np.array(increase_values) - np.array(labels))
        cluster_score1 = sum(-3 * pow(x, 2) + 3.5 * x)
        x = abs((1 - np.array(increase_values)) - np.array(labels))
        cluster_score2 = sum(-3 * pow(x, 2) + 3.5 * x)

        # for i, value in enumerate(increase_values):
        #    if labels[i] == 1 and value == 1:

        if cluster_score1 < cluster_score2:
            eps_.append(cluster_score1)
            labels_tmp = np.array(labels)
        else:
            eps_.append(cluster_score2)
            labels_tmp = 1 - np.array(labels)
        """

        labels_tmp = np.array(labels)

        all_labels.append(copy.deepcopy(labels_tmp.tolist()))
        all_labels_original.append(copy.deepcopy(labels_original))

        # print(f'eps = {i} min_samples = {j}\n{min(cluster_score1, cluster_score2)}\n')
        cnt += 1
        print(f"========= progressing {100 * cnt / (len(eps_cand) * len(min_sam_cand))}% =========")
        print(f'eps = {eps_val} min_samples = {min_sam_val}\n{labels_tmp}')

    eps_sam.append(eps_)

print(f'eps_sam is {eps_sam}')

write_wb = Workbook()
write_ws = write_wb.create_sheet('table')
write_ws.append(['eps\sam'] + min_sam_cand)
for i, eps in enumerate(eps_sam):
    write_ws.append([eps_cand[i]] + eps)

cnt2 = 0
precision_value = []
accuracy_list = []
precision_value_train = []
accuracy_list_train = []
precision_value_new = []
accuracy_list_new = []
for i, eps_val in enumerate(eps_cand):
    for j, min_sam_val in enumerate(min_sam_cand):
        write_ws = write_wb.create_sheet(f'{eps_val}-{min_sam_val}')
        accuracy = 100 * np.sum(np.equal(increase_values, all_labels[cnt2])) / len(increase_values)
        accuracy_list.append(accuracy)
        write_ws.append(['accuracy', accuracy])
        write_ws.append(['all_cluster_label', str(all_cluster_label)])
        write_ws.append(['gt', 'prediction', 'pred_original'])
        for i, increase_value in enumerate(increase_values):
            write_ws.append([increase_value, all_labels[cnt2][i], all_labels_original[cnt2][i]])

        print(f'\neps = {eps_val} min_samples = {min_sam_val}')


        #print(f'accuracy is {100 * np.sum(np.equal(increase_values, all_labels[cnt2])) / len(increase_values)}')

        prediction = np.array(all_labels[cnt2])
        precision = np.where(prediction == 1)
        #print(f'precision is {precision}')
        precision_cnt = 0
        for precision_index in precision[0]:
            #print(f'precision_index is {precision_index}')
            if increase_values[precision_index] == 1:
                precision_cnt += 1
        if precision_cnt != 0:
            precision_value.append(100 * precision_cnt / len(precision[0].tolist()))
            write_ws.append([100 * precision_cnt / len(precision[0].tolist())])
        else:
            precision_value.append(0)
            write_ws.append([0])


        prediction_train = np.array(all_labels[cnt2][:-test_size])
        precision_train = np.where(prediction_train == 1)
        precision_train_cnt = 0
        for precision_index in precision_train[0]:
            # print(f'precision_index is {precision_index}')
            if increase_values[:-test_size][precision_index] == 1:
                precision_train_cnt += 1
        if precision_train_cnt != 0:
            precision_value_train.append(100 * precision_train_cnt / len(precision_train[0].tolist()))
            write_ws.append([100 * precision_train_cnt / len(precision_train[0].tolist())])
        else:
            precision_value_train.append(0)
            write_ws.append([0])
        print(f'precision_value_train is {precision_value_train}')



        prediction_new = np.array(all_labels[cnt2][-test_size:])
        precision_new = np.where(prediction_new == 1)
        precision_new_cnt = 0
        for precision_index in precision_new[0]:
            # print(f'precision_index is {precision_index}')
            if increase_values[-test_size:][precision_index] == 1:
                precision_new_cnt += 1
        if precision_new_cnt != 0:
            precision_value_new.append(100 * precision_new_cnt / len(precision_new[0].tolist()))
            write_ws.append([100 * precision_new_cnt / len(precision_new[0].tolist())])
        else:
            precision_value_new.append(0)
            write_ws.append([0])
        print(f'precision_value_new is {precision_value_new}')




        cnt2 += 1

write_ws = write_wb.create_sheet('precision')
precision_write_cnt = 0
write_ws.append(['eps\sam'] + min_sam_cand)
for i, eps_val in enumerate(eps_cand):
    precision_write = []
    for j, min_sam_val in enumerate(min_sam_cand):
        precision_write.append(precision_value[precision_write_cnt])
        precision_write_cnt += 1
    write_ws.append([eps_val] + precision_write)


write_ws = write_wb.create_sheet('precision_train')
precision_train_write_cnt = 0
write_ws.append(['eps\sam'] + min_sam_cand)
for i, eps_val in enumerate(eps_cand):
    precision_train_write = []
    for j, min_sam_val in enumerate(min_sam_cand):
        precision_train_write.append(precision_value_train[precision_train_write_cnt])
        precision_train_write_cnt += 1
    write_ws.append([eps_val] + precision_train_write)


write_ws = write_wb.create_sheet('precision_new')
precision_new_write_cnt = 0
write_ws.append(['eps\sam'] + min_sam_cand)
for i, eps_val in enumerate(eps_cand):
    precision_new_write = []
    for j, min_sam_val in enumerate(min_sam_cand):
        precision_new_write.append(precision_value_new[precision_new_write_cnt])
        precision_new_write_cnt += 1
    write_ws.append([eps_val] + precision_new_write)

write_ws = write_wb.create_sheet('accuracy')
accuracy_write_cnt = 0
write_ws.append(['eps\sam'] + min_sam_cand)
for i, eps_val in enumerate(eps_cand):
    accuracy_write = []
    for j, min_sam_val in enumerate(min_sam_cand):
        accuracy_write.append(accuracy_list[accuracy_write_cnt])
        accuracy_write_cnt += 1
    write_ws.append([eps_val] + accuracy_write)

write_wb.save('eps_sam.xlsx')
print(f'precision_value is {precision_value}')




